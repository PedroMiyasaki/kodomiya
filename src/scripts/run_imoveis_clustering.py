def main() -> None:
    import pandas as pd
    import numpy as np
    from scipy import stats
    import duckdb
    from datetime import datetime
    from sklearn.cluster import KMeans
    from sklearn.metrics import silhouette_score
    from sklearn.preprocessing import MinMaxScaler
    import warnings
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.borders import Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows


    warnings.filterwarnings("ignore")


    def find_best_n_of_clusters(dataframe, columns, cluster_range):
        """
        Perform clustering using KMeans and find the best number of clusters based on silhouette score.

        Parameters:
        - dataframe: DataFrame containing the dataframe for clustering.
        - columns: columns in the DataFrame to use for clustering.
        - cluster_range: Range of cluster numbers to consider.

        Returns:
        - best_clusters: Number of clusters that yield the highest silhouette score.
        - cluster_labels: Cluster labels assigned by KMeans.
        """
        features = dataframe[columns]
        mms = MinMaxScaler()
        features = mms.fit_transform(dataframe[columns])

        best_score = -1
        best_clusters = 0
        cluster_labels = None
        n_clusters_history = {"n_clusters": [], "score": []}

        for n_clusters in cluster_range:
            if n_clusters >= len(features):
                break

            kmeans = KMeans(n_clusters=n_clusters, random_state=32, n_init="auto")
            labels = kmeans.fit_predict(features)
            score = silhouette_score(features, labels)

            n_clusters_history["n_clusters"].append(n_clusters)
            n_clusters_history["score"].append(score)
            
            if score > best_score:
                best_score = score
                best_clusters = n_clusters
                cluster_labels = labels

        print(f"Best Number of Clusters: {best_clusters}")
        print(f"Silhouette Score: {best_score}")

        return best_clusters, cluster_labels, n_clusters_history


    # Database connection
    con = duckdb.connect(database="../db/trading_properties_db.duckdb")

    # Load data from different sources
    df_chaves_na_mao = con.execute("""
        SELECT 
            * 
        FROM trading_properties_db.chaves_na_mao_schema.chaves_na_mao_register
    """).fetch_df()
    df_chaves_na_mao["fonte_dado"] = "Chaves na Mão"

    df_zap_imoveis = con.execute("""
        SELECT 
            * 
        FROM trading_properties_db.zap_imoveis_schema.zap_imoveis_register
    """).fetch_df()
    df_zap_imoveis["fonte_dado"] = "Zap Imóveis"

    df_viva_real = con.execute("""
        SELECT 
            * 
        FROM trading_properties_db.viva_real_schema.viva_real_register
    """).fetch_df()
    df_viva_real["fonte_dado"] = "Viva Real"

    con.close()
    
    # Concatenate all dataframes
    df = pd.concat([df_chaves_na_mao, df_zap_imoveis, df_viva_real], ignore_index=True)

    # Remove duplicate property IDs keeping the latest
    df = df.sort_values("datahora").drop_duplicates("id", keep="last")

    # Create unique ID for latitude and longitude
    df["lat_long_id"] = df["latitude"].apply(str) + "-" + df["longitude"].apply(str)

    # Count occurrences of each lat-long pair
    df["count_lat_log_id"] = df.groupby("lat_long_id")["id"].transform("count")

    # Set duplicate coordinates to null
    df.loc[df["count_lat_log_id"] > 1, "latitude"] = None
    df.loc[df["count_lat_log_id"] > 1, "longitude"] = None
    
    # Remove rows with null price or size
    df = df.loc[(~df["tamanho"].isna()) & (~df["preco"].isna())]

    # Remove outliers (3 Z-scores) for price and size
    df = df[(np.abs(stats.zscore(df[["tamanho", "preco"]])) < 3).all(axis=1)]

    # Run general clustering process
    _, cluster_labels, _ = find_best_n_of_clusters(
        df, 
        ["tamanho", "n_banheiros", "n_quartos", "n_garagem"], 
        range(2, 100)
    )

    df["cluster_geral"] = cluster_labels

    # Initialize list for neighborhood dataframes
    bairro_dataframes = []

    # Run clustering process by neighborhood
    for bairro in df["bairro"].dropna().unique():
        df_bairro = df.loc[df["bairro"] == bairro].copy()
        
        _, cluster_labels, _ = find_best_n_of_clusters(
            df_bairro, 
            ["tamanho", "n_banheiros", "n_quartos", "n_garagem"], 
            range(2, 30)
        )

        df_bairro["cluster_bairro"] = cluster_labels
        bairro_dataframes.append(df_bairro)

    bairro_dataframe = pd.concat(bairro_dataframes, ignore_index=True)

    # Merge neighborhood clusters back to original dataframe
    df = df.merge(bairro_dataframe[["id", "cluster_bairro"]], on="id", how="left")

    # Set cluster as null for "not found" neighborhoods
    df.loc[df["bairro"] == "não encontrado", "cluster_bairro"] = None

    # Calculate average prices by cluster
    df["preco_medio_cluster_geral"] = df.groupby("cluster_geral")["preco"].transform("mean")
    df["preco_medio_cluster_bairro"] = df.groupby(["bairro", "cluster_bairro"])["preco"].transform("mean")

    # Calculate price difference percentage from cluster average
    df["dif_pct_cluster_geral"] = (df["preco"] - df["preco_medio_cluster_geral"]) / df["preco_medio_cluster_geral"]
    df["dif_pct_cluster_bairro"] = (df["preco"] - df["preco_medio_cluster_bairro"]) / df["preco_medio_cluster_bairro"]

    # Calculate standard deviation within clusters
    df["std_cluster_geral"] = df.groupby("cluster_geral")["preco"].transform("std")
    df["std_cluster_bairro"] = df.groupby(["bairro", "cluster_bairro"])["preco"].transform("std")

    # Calculate statistical "Discount" (Z-score)
    df["z_value_cluster_geral"] = (df["preco"] - df["preco_medio_cluster_geral"]) / df["std_cluster_geral"] 
    df["z_value_cluster_bairro"] = (df["preco"] - df["preco_medio_cluster_bairro"]) / df["std_cluster_bairro"] 

    # Remove timezone from datetime
    df["datahora"] = df["datahora"].dt.tz_localize(None)

    # Select required columns
    df = df[[
        "id", "datahora", "preco", "tamanho", "n_quartos", "n_banheiros",
        "n_garagem", "rua", "bairro", "cidade", 
        "latitude", "longitude", "cluster_geral", 
        "cluster_bairro", "preco_medio_cluster_geral",
        "preco_medio_cluster_bairro", "dif_pct_cluster_geral",
        "dif_pct_cluster_bairro", "z_value_cluster_geral", 
        "z_value_cluster_bairro", "fonte_dado"
    ]]

    # Rename columns for output
    df.columns = [
        "ID Imóvel", "DataHora", "Preço", "Tamanho (m²)", 
        "Qtd Quartos (#)", "Qtd Banheiros (#)", "Qtd Vagas Garagem (#)",
        "Rua", "Bairro", "Cidade", "Latitude", "Longitude", 
        "Cluster Geral", "Cluster Bairro",
        "Preço médio do Imóvel (Cluster Geral)", 
        "Preço médio do Imóvel (Cluster Bairro)",
        "Diferença % preço imóvel (Média Cluster Geral)",
        "Diferença % preço imóvel (Média Cluster Bairro)",
        "Z-Valor preço imóvel (Cluster Geral)",
        "Z-Valor preço imóvel (Cluster Bairro)",
        "Site Anúncio"
    ]
    
    # Title case for neighborhood and city names
    df["Bairro"] = df["Bairro"].str.title()
    df["Cidade"] = df["Cidade"].str.title()

    # Create cluster description dataframes
    descritivo_clusters_bairro = df[[
        "Bairro", "Cluster Bairro", "Preço", "Tamanho (m²)", 
        "Qtd Quartos (#)", "Qtd Banheiros (#)", "Qtd Vagas Garagem (#)"
    ]]
    
    descritivo_clusters = df[[
        "Cluster Geral", "Preço", "Tamanho (m²)", 
        "Qtd Quartos (#)", "Qtd Banheiros (#)", "Qtd Vagas Garagem (#)"
    ]]

    # Group dataframes by mean
    descritivo_clusters_bairro = descritivo_clusters_bairro.groupby(["Bairro", "Cluster Bairro"]).mean()
    descritivo_clusters = descritivo_clusters.groupby("Cluster Geral").mean()

    # Rename columns to indicate averages
    descritivo_clusters_bairro = descritivo_clusters_bairro.rename(
        columns={c: "Média " + c for c in descritivo_clusters_bairro.columns}
    )
    descritivo_clusters = descritivo_clusters.rename(
        columns={c: "Média " + c for c in descritivo_clusters.columns}
    )

    # Round numeric values
    descritivo_clusters_bairro = descritivo_clusters_bairro.round(2)
    descritivo_clusters = descritivo_clusters.round(2)

    # Sort by size
    descritivo_clusters_bairro = descritivo_clusters_bairro.sort_values(["Bairro", "Média Tamanho (m²)"])
    descritivo_clusters = descritivo_clusters.sort_values("Média Tamanho (m²)")

    # Reset index
    descritivo_clusters_bairro = descritivo_clusters_bairro.reset_index()
    descritivo_clusters = descritivo_clusters.reset_index()

    # Define border styles
    medium_border_style = Border(
        left=Side(style="medium"), 
        right=Side(style="medium"), 
        top=Side(style="medium"), 
        bottom=Side(style="medium")
    )

    thin_border_style = Border(
        left=Side(style="dashed"), 
        right=Side(style="dashed"), 
        top=Side(style="dashed"), 
        bottom=Side(style="dashed")
    )

    # Create workbook
    workbook = openpyxl.Workbook()

    # Create sheets for data storage
    worksheet_clusterizacao = workbook.create_sheet("Property Clustering")
    worksheet_descritivo_cluster_bairro = workbook.create_sheet("Neighborhood Cluster Averages")
    worksheet_descritivo_cluster_geral = workbook.create_sheet("General Cluster Averages")

    # Define heading style
    default_heading_style = openpyxl.styles.NamedStyle(name="default_heading_style")
    default_heading_style.font = Font(bold=True, color="000000")
    default_heading_style.fill = PatternFill(start_color="B0D0FF", end_color="B0D0FF", fill_type="solid")

    # Define value cell style
    default_style = openpyxl.styles.NamedStyle(name="value_style")
    default_style.fill = PatternFill(start_color="FEFEFE", end_color="FEFEFE", fill_type="solid")

    # Copy dataframes to avoid memory issues
    worksheet_clusterizacao_data = df.copy()
    worksheet_descritivo_cluster_bairro_data = descritivo_clusters_bairro.copy()
    worksheet_descritivo_cluster_geral_data = descritivo_clusters.copy()

    # Load pandas dataframe rows
    worksheet_clusterizacao_rows = dataframe_to_rows(worksheet_clusterizacao_data, index=False, header=True)
    worksheet_descritivo_cluster_bairro_rows = dataframe_to_rows(worksheet_descritivo_cluster_bairro_data, index=False, header=True)
    worksheet_descritivo_cluster_geral_rows = dataframe_to_rows(worksheet_descritivo_cluster_geral_data, index=False, header=True)

    # Iterate over rows, sheets, and data formats
    for rows, worksheet, list_of_formats in [
        (worksheet_clusterizacao_rows, worksheet_clusterizacao, [
            'General', 'mm-dd-yy', 'R$   #,##0.00', '0', '0', '0', 
            '0', 'General', 'General', '0.00', '0.00', 
            'General', '0', '0', 'R$   #,##0.00', 
            'R$   #,##0.00', '0.00%', '0.00%', '0.00', '0.00', 'General'
        ]), 
        (worksheet_descritivo_cluster_bairro_rows, worksheet_descritivo_cluster_bairro, [
            'General', '0', 'R$   #,##0.00', '0.00', '0.00', '0.00', '0.00',
        ]), 
        (worksheet_descritivo_cluster_geral_rows, worksheet_descritivo_cluster_geral, [
            '0', 'R$   #,##0.00', '0.00', '0.00', '0.00', '0.00',
        ])
    ]:
        # Iterate all rows of current worksheet and set values
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Apply heading style
        for row in worksheet.iter_rows(min_row=0, max_row=1):
            for cell in row:
                cell.style = default_heading_style
                cell.border = medium_border_style

        # Apply row styles
        for r_idx, row in enumerate(worksheet.iter_rows(min_row=2)):
            for cell_idx, cell in enumerate(row):
                cell.style = default_style
                cell.border = thin_border_style
                cell.number_format = list_of_formats[cell_idx]

    # Remove default sheet
    del workbook["Sheet"]

    # Save formatted version
    workbook.save(f"../analytics/Property_Clustering_{datetime.today().strftime('%Y_%m_%d')}.xlsx")

    return None


if __name__ == '__main__':
    main()
